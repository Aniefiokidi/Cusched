import io
import csv
from models.models import TimetableEntry

DAYS = ["MON", "TUE", "WED", "THU", "FRI"]
TIMES = ["08:00", "09:00", "10:00", "11:00", "12:00", "13:00", "14:00", "15:00", "16:00"]


def get_all_entries():
    entries = TimetableEntry.query.all()
    return [e.to_dict() for e in entries]


def export_csv():
    entries = get_all_entries()
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[
        "course_code", "course_title", "lecturer", "room",
        "student_group", "day", "start_time", "end_time"
    ])
    writer.writeheader()
    for e in entries:
        writer.writerow({k: e.get(k, "") for k in writer.fieldnames})
    return output.getvalue()


def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    entries = get_all_entries()
    wb = Workbook()
    ws = wb.active
    ws.title = "Timetable"

    headers = ["Course Code", "Course Title", "Lecturer", "Room", "Student Group", "Day", "Start", "End"]
    header_fill = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin

    for row, e in enumerate(entries, 2):
        vals = [e.get("course_code"), e.get("course_title"), e.get("lecturer"),
                e.get("room"), e.get("student_group"), e.get("day"),
                e.get("start_time"), e.get("end_time")]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin
            cell.alignment = Alignment(horizontal="left")

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col) + 4
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 35)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    entries = get_all_entries()
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                            rightMargin=1*cm, leftMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Heading1"],
                                  fontSize=16, alignment=TA_CENTER,
                                  textColor=colors.HexColor("#1A3C5E"))
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"],
                                fontSize=10, alignment=TA_CENTER,
                                textColor=colors.HexColor("#6B7280"))

    story = [
        Paragraph("CovenantSched — Generated Timetable", title_style),
        Paragraph("Covenant University — Department of Computer and Information Sciences", sub_style),
        Spacer(1, 0.5*cm),
    ]

    header = ["Course", "Title", "Lecturer", "Room", "Group", "Day", "Start", "End"]
    table_data = [header]
    for e in entries:
        table_data.append([
            e.get("course_code", ""),
            e.get("course_title", "")[:30],
            e.get("lecturer", ""),
            e.get("room", ""),
            e.get("student_group", ""),
            e.get("day", ""),
            e.get("start_time", ""),
            e.get("end_time", ""),
        ])

    navy = colors.HexColor("#1A3C5E")
    light = colors.HexColor("#F0F4F8")
    tbl = Table(table_data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, light]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CDD5E0")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)
    doc.build(story)
    output.seek(0)
    return output
